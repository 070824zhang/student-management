import openpyxl
from openpyxl import load_workbook
import os

class Student:
    def __init__(self,name,gender,age,address):
        self.name = name
        self.gander = gender
        self.age = age
        self.address = address

#定义学生id并设置id自增
def get_next_id():
    #如果Excel文件不存在返回1
    if not os.path.exists('students.xlsx'):
        return 1
    #打开现有Excel文件
    wb = load_workbook('students.xlsx')
    ws = wb.active
    max_id = 0
    student_read = list(ws.iter_rows(min_row=2, values_only=True))#从第二行开始，跳过表头
    for student_data in student_read:            #student_data是一个元组，例如（1，‘张三’，‘男’，20，’北京‘）
        if not student_data[0]:
            continue
        student_id = int(student_data[0])
        if student_id > max_id:  # 实现增加一条学生信息，id自增
            max_id = student_id
        return max_id + 1


#定义函数，用来增加学生信息

def add_student():
    import os
    print("当前工作目录,", os.getcwd())
    next_id = get_next_id()   #调用get_next_id函数，自动为录入的学生信息设置id
    if not os.path.exists('students.xlsx'):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['id','姓名','性别','年龄','地址'])
        wb.save('students.xlsx')
        print("文件创建成功")
    else:
        print("文件已经存在")
        #打开现有文件准备追加
    from openpyxl import load_workbook
    wb = load_workbook('students.xlsx')
    ws = wb.active
    while True:
        print("请录入学生信息")
        name = input("请输入学生姓名：")
        gender = input("请输入学生性别")
        age = input("请输入学生年龄：")
        address = input("请输入学生家庭地址：")
        #追加一行
        ws.append([next_id,name,gender,age,address])
        wb.save('students.xlsx')
        print(f"学生{name}添加成功，ID={next_id}")
        next_id += 1
        print("当前学生信息录入完毕，是否继续录入？输入1继续录入，否则退出。")
        choose = input("请确认是否继续录入：")
        if choose != "1":
            break




def del_student():
    try:
        sid = int(input("请输入要删除的学生id："))
    except ValueError:
        print("请输入数字")
        return
    if not os.path.exists('students.xlsx'):
        print("文件不存在")
        return
    wb = load_workbook()
    ws = wb.active
    row_to_del = None
    for row in ws.iter_rows(min_row=2):
        if row[0].value == sid:
            row_to_del = row[0].row
            break
        if row_to_del is None:
            print("未找到该学生")
            return
    ws.delete_rows(row_to_del)
    wb.save('students.xlsx')
    print("删除成功！")



def find_student_by_id(sid):
    if not os.path.exists('students.xlsx'):
        return None
    wb = load_workbook('students.xlsx')
    ws = wb.active
    #遍历从第二行开始
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == sid:
            return row
    return None




def updata_student():
    try:
        sid = int(input("请输入要修改的学生id："))
    except ValueError:
        print("请输入数字")
        return
    if not os.path.exists('students.xlsx'):
        print("文件不存在")
        return

    #打开Excel，查找目标行
    wb = load_workbook('students.xlsx')
    ws = wb.active
    target_row = None
    for row in ws.iter_rows(min_row=2):
        if row[0].value == sid:
            target_row = row[0].row
            break
        if target_row is None:
            print("未找到该学生")
            return
        name_cell = ws.cell(row=target_row, column=2)
        gender_cell = ws.cell(row=target_row, column=3)
        age_cell = ws.cell(row=target_row, column=4)
        address_cell = ws.cell(row=target_row, column=5)
        print(f"当前信息：姓名：{name_cell.value},性别：{gender_cell.value},年龄：{age_cell.value},地址：{address_cell.value}")
        choose = input("请输入要修改的字段：1姓名 2性别 3年龄 4地址\n")
        if choose == "1":
            name_cell.value = input("新姓名：")
        elif choose == "2":
            gender_cell.value = input("新性别：")
        elif choose == "3":
            age_cell.value = input("新年龄：")
        elif choose == "4":
            address_cell.value = input("新地址：")
        else:
            print("无效选择")
            return
        wb.save('students.xlsx')
        print("修改成功")





def inquire_student():
    try:
        sid = int(input("请输入要查询学生的id："))
        result = find_student_by_id(sid)
        if result is None:
            print("未查询到该学生信息")
            return
        print(f"学生信息：id={result[0]},姓名={result[1]},性别={result[2]},年龄={result[3]},住址={result[4]}")
    except ValueError:
        print("请输入数字")
        return








def main():
    while True:
        print("欢迎进入学生信息管理系统，请选择操作，输入1增加学生信息,输入2删除学生信息，输入3修改学生信息，输入4查询学生信息")
        num = int(input("请输入数字："))
        if num == 1:
            add_student()
        elif num == 2:
            del_student()
        elif num == 3:
            updata_student()
        elif num == 4:
            inquire_student()
        else:
            print("您已退出学生信息管理系统")
            break








if __name__ == '__main__':
    main()